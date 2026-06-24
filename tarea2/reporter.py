import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
from openpyxl.utils import get_column_letter

# --- 1. OPENPYXL HELPERS (Funciones de navegación) ---

def _encontrar_columna_mes(sheet, mes_objetivo: int) -> str:
    """
    Busca la columna correspondiente al mes objetivo (ej: 8 para Agosto).
    Asume que las fechas están en la Fila 3.
    
    Args:
        sheet: Objeto de la hoja de openpyxl.
        mes_objetivo (int): Número del mes (1-12).
    
    Returns:
        str: Letra de la columna (ej: 'J').
    """
    fila_fechas = 3 

    header_row = next(sheet.iter_rows(min_row=fila_fechas, max_row=fila_fechas, values_only=False))

    for cell_obj in header_row:
        if isinstance(cell_obj.value, datetime) and cell_obj.value.month == mes_objetivo:
            return get_column_letter(cell_obj.column)

    month_token = f"/{str(mes_objetivo).zfill(2)}/"
    for cell_obj in header_row:
        if isinstance(cell_obj.value, str) and month_token in cell_obj.value:
            return get_column_letter(cell_obj.column)
                
    raise ValueError(f"❌ No se encontró la columna para el mes {mes_objetivo} en la fila {fila_fechas}.")

def _encontrar_fila_por_texto(sheet, search_text: str) -> int:
    """
    Busca un texto específico (ej: nombre del Roll Up) en la hoja y devuelve su número de fila.
    """
    # Itera sobre las primeras 5 columnas para evitar buscar en celdas de datos o fórmulas
    for row in sheet.iter_rows(min_col=2, max_col=5): 
        for cell in row:
            # strip() y lower() para hacer la búsqueda más robusta
            if cell.value and str(cell.value).strip().lower() == search_text.strip().lower():
                return cell.row
    return None

# --- 2. CORE REPORTER FUNCTIONS ---

def actualizar_pl_excel_local(ruta_template: Path, ruta_salida: Path, nii: float,
                              taxes: float, gastos_rollup: pd.Series, mes_objetivo: int):
    """
    Abre el template P&L, inserta los valores calculados de NII, Gastos (Roll Up) y Taxes, y guarda el resultado.
    """
    print(f"📝 Actualizando P&L en: {ruta_salida.parent.name}")
    
    try:
        wb = openpyxl.load_workbook(ruta_template)
    except Exception as e:
        raise FileNotFoundError(f"❌ Error al cargar template: {ruta_template}. {e}")
        
    sheet = wb['P&L ARG'] # Usar nombre de hoja fijo
    
    # 1. Encontrar la columna de escritura (ej: 'J' para Agosto)
    columna = _encontrar_columna_mes(sheet, mes_objetivo)
    print(f"   Columna de escritura identificada: {columna}")
    
    # 2. Mapeo fijo de filas (Basado en el layout de tu P&L original)
    # Nota: Estos son los puntos de partida fijos que identificaste en tu notebook
    ROW_NII = 4       
    ROW_TAXES = 44    
    
    # 3. Escribir NII y Taxes (totales simples)
    sheet[f"{columna}{ROW_NII}"] = nii * -1 # Multiplicado por -1 según la lógica original de tu notebook
    sheet[f"{columna}{ROW_TAXES}"] = taxes
    
    # 4. Escribir Gastos (Roll Up)
    for roll_up_name, valor in gastos_rollup.items():
        fila = _encontrar_fila_por_texto(sheet, roll_up_name)
        
        if fila:
            sheet[f"{columna}{fila}"] = round(valor, 2)
        else:
            print(f"⚠️ Advertencia: No se encontró la fila para el Roll Up: '{roll_up_name}'.")

    # 5. Guardar el archivo final
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta_salida)
    print(f"   💾 Guardado en: {ruta_salida.name}")


def guardar_reporte_um_local(df: pd.DataFrame, ruta_salida: Path):
    """
    Guarda el DataFrame Reporte UM en formato Excel.
    """
    print(f"📊 Guardando Reporte UM...")
    
    # Asegurar que el directorio de salida exista
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    
    # Escribir el DataFrame al disco (sin encabezado ni índice de fila)
    df.to_excel(ruta_salida, sheet_name='ReportUM', header=False, index=True)
    print(f"   💾 Guardado en: {ruta_salida.name}")